"""Test fixtures: build xlsx and csv inputs into tmp_path."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def csv_file(tmp_path: Path) -> Path:
    p = tmp_path / "contacts.csv"
    p.write_text(
        "Email,First Name,Phone,Joined\n"
        "a@x.com,Alice,(555) 111-2222,2024-01-15\n"
        "b@x.com,Bob,555.333.4444,2024-02-20\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture
def xlsx_file(tmp_path: Path) -> Path:
    p = tmp_path / "sales.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    jan = wb.create_sheet("Jan")
    jan.append(["Order ID", "Order Date", "Amount", "Notes"])
    jan.append([101, "2024-01-05", "$1,234.50", "first"])
    jan.append([102, "2024-01-15", "$200.00", None])
    feb = wb.create_sheet("Feb")
    feb.append(["Order ID", "Order Date", "Amount", "Notes"])
    feb.append([201, "2024-02-03", "$99.99", "second"])
    wb.save(p)
    return p
